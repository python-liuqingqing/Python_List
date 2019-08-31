import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import re
import urllib
import random
import threading
import mysql.connector
from selenium import webdriver


i=1
l=2
def url_pa(i,l):
    wb = load_workbook('11.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    lou_name = []
    for rows in range(i, l):
        key_word = sheet.cell(row=rows + 1, column=2).value
        lou_name.append(key_word)
        print(key_word, rows)
        print('正在搜索，请稍后')

        driver = webdriver.Firefox()  # 打开火狐浏览器
        driver.get('https://www.crunchbase.com/organization/kyc-exchange-net#section-overview')  # 打开界面

        # url = r'http://www.qichacha.com/search?key={}#p:{}&'.format(key_word, 1)
        # s1 = craw("https://www.qichacha.com/search?key="+key_word,key_word)
        # url_company = "http://www.qichacha.com" + s1
        # crawler_company(url_company, key_word)
        #
    i = i+1
    l = l+1
url_pa(i,l)