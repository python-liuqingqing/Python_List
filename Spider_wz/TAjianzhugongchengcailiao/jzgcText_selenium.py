# coding:utf-8
from selenium import webdriver
from openpyxl import load_workbook
import random
import threading
import time
import mysql.connector
from bs4 import BeautifulSoup
from lxml import etree
import os
jsonData = []
def main(key_word, rows,i):
    driver=webdriver.Firefox()
    driver.get(key_word)
    time.sleep(10)
    # 获取页面html
    file = open('D:\TA\labels'+str(i)+'.txt', 'w')
    file.close()
    html = driver.page_source
    fos = open('D:\TA\labels'+str(i)+'.txt', "w", encoding="utf-8")
    fos.write(html)
    fos.close()
    driver.quit()
i=1
l=2
def url_pa(i,l):
    # excel文件名称
    wb = load_workbook('ta.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    lou_name = []
    for rows in range(i, l):
        key_word = sheet.cell(row=rows + 1, column=2).value
        lou_name.append(key_word)
        print(key_word, rows)
        print('正在搜索，请稍后')
        main(key_word, rows,i)
    i = i+1
    l = l+1
    threading.Timer(5, url_pa, [i, l]).start()
if __name__ == '__main__':
    threading.Timer(5,url_pa, [i,l]).start()