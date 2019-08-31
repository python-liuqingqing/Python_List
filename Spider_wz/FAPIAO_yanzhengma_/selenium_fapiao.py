# coding=utf-8
import time
import mysql.connector
import pytesseract
from PIL import Image
from selenium import webdriver
from openpyxl import load_workbook
import threading

driver = webdriver.Firefox()  # 打开火狐浏览器
# driver=webdriver.PhantomJS(executable_path="C:\Python36\Scripts\phantomjs.exe")
driver.get('http://fpcx.hebds.gov.cn/fpxx/fpxxcx.jsp')  # 打开界面

def ff(i,a1,a2,a3):
    js = 'window.open("http://fpcx.hebds.gov.cn/fpxx/fpxxcx.jsp");'
    driver.execute_script(js)
    handles = driver.window_handles  # 获取当前窗口句柄集合（列表类型）
    for handle in handles:  # 切换窗口（切换到搜狗）
        if handle != driver.current_window_handle:
            driver.switch_to_window(handle)
    time.sleep(3)

    driver.find_element_by_id('fpdm').send_keys(a1)
    driver.find_element_by_id('fphm').send_keys(a2)
    driver.find_element_by_id('dwmc').send_keys(a3)
    time.sleep(2)
    driver.find_element_by_xpath('/html/body/div/div[4]/table/tbody/tr/td[2]/a').click()
    time.sleep(2)
    driver.maximize_window()
    driver.save_screenshot('D:/YZM/s'+str(i)+'.png')  # 截取当前网页，该网页有我们需要的验证码
    imgelement = driver.find_element_by_id('cfd')
    location = imgelement.location  # 获取验证码x,y轴坐标
    size = imgelement.size  # 获取验证码的长宽
    rangle = (int(location['x']), int(location['y']), int(location['x'] + size['width']),int(location['y'] + size['height']))  # 写成我们需要截取的位置坐标
    print(int(location['x']), int(location['y']), int(location['x'] + size['width']),int(location['y'] + size['height']))
    i = Image.open('D:/YZM/s'+str(i)+'.png')  # 打开截图
    result = i.crop(rangle)  # 使用Image的crop函数，从截图中再次截取我们需要的区域
    result.save('D:/YZM/1.png')
    # driver.switch_to.frame(driver.find_element_by_id('cfd'))  # 切入
    # img=driver.find_element_by_xpath('/html/body/img').get_attribute('src')
    # urllib.request.urlretrieve(img,'D:/YZM/'+str(i)+'.jpg')
    # driver.switch_to_default_content()   # 切出
    # # 2
    image = Image.open('D:/YZM/1.png')
    result = pytesseract.image_to_text(image)
    print(result)
    driver.find_element_by_id('validateCode').send_keys(result)
    driver.find_element_by_id('cxButton').click()
    time.sleep(10)
    try:
        #收款单位
        skdw = driver.find_element_by_id('dw').get_attribute('value') #if driver.find_element_by_id('dw').get_attribute('value') !="" else a3
        #主管税务机关
        zgswjg = driver.find_element_by_id('swjg').get_attribute('value')
        #发票代码
        fpdm = driver.find_element_by_id('cxfpdm').get_attribute('value') #if driver.find_element_by_id('cxfpdm').get_attribute('value') !="" else a1
        #发票号码
        fphm = driver.find_element_by_id('cxfphm').get_attribute('value') #if driver.find_element_by_id('cxfphm').get_attribute('value') !="" else a2
        #购票日期
        gprq = driver.find_element_by_id('kprq').get_attribute('value')
        #开票金额
        kpje = driver.find_element_by_id('kpje').get_attribute('value')
        #发票状态
        fpzt = driver.find_element_by_id('fpzt').get_attribute('value')
        #查询次数
        cxcs = driver.find_element_by_id('cxcs').get_attribute('value')
        #付款单位名称
        fkdwmc = driver.find_element_by_id('fkdwmc').get_attribute('value')
        #付款单位识别号
        fkdwsbh     = driver.find_element_by_id('fkdwsbh').get_attribute('value')
        #非定额发票
        fdefp = driver.find_element_by_id('me').get_attribute('value')
        #
        print(skdw,zgswjg,fpdm,fphm,gprq,kpje,fpzt,cxcs,fkdwmc,fkdwsbh,fdefp)
        jh = [{
            skdw:skdw,
            zgswjg:zgswjg,
            fpdm:fpdm,
            fphm:fphm,
            gprq:gprq,
            kpje:kpje,
            fpzt:fpzt,
            cxcs:cxcs,
            fkdwmc:fkdwmc,
            fkdwsbh:fkdwsbh,
            fdefp:fdefp
        }]

        # import json
        # json = json.dumps(jh)

        if(skdw!=""):
            db = mysql.connector.connect(user='root', password='123456', host='127.0.0.1', database='fapiao_hb_20181001')
            cursor = db.cursor()

            insert_color = ("INSERT INTO fapiao_hb(skdw,zgswjg,fpdm,fphm,gprq,kpje,fpzt,cxcs,fkdwmc,fkdwsbh,fdefp,time)" "VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)")
            data_color = (skdw,zgswjg,fpdm,fphm,gprq,kpje,fpzt,cxcs,fkdwmc,fkdwsbh,fdefp,time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time())))
            cursor.execute(insert_color, data_color)
            db.commit()
            #
            html=driver.page_source
            # print(html)
            driver.close()
            driver.switch_to_window(handles[0])
        return 'a'
    except Exception as e:
        driver.close()
        driver.switch_to_window(handles[0])
        return 'b'


i = 1
l = 2
# excel文件名称
def main_(i,l):
    wb = load_workbook('123.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    lou_name = []
    for rows in range(i, l):
        a1 = sheet.cell(row=rows+1, column=1).value
        a2 = sheet.cell(row=rows+1, column=2).value
        a3 = sheet.cell(row=rows+1, column=3).value
        print(i,a1,a2,a3)

        # lou_name.append(key_word)
        # print(key_word, rows)
        if(ff(i,a1,a2,a3)=='a'):
            i = i+1
            l = l+1
        elif(ff(i,a1,a2,a3)=="b"):
            i = i
            l = l
        threading.Timer(5, main_, [i, l]).start()
if __name__ == '__main__':
    threading.Timer(5,main_, [i,l]).start()