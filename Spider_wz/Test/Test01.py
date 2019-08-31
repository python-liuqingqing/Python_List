# -*- coding: utf-8 -*-
import requests
import json
from openpyxl import load_workbook
from openpyxl import workbook
import threading

requests.packages.urllib3.disable_warnings()
wb1 = workbook.Workbook()  # 创建Excel对象
ws1 = wb1.active  # 获取当前正在操作的表对象

def craw(key_word):
    url = 'https://app.gsxt.gov.cn/gsxt/cn/gov/saic/web/controller/PrimaryInfoIndexAppController/search?page=1'
    body = json.loads('{"searchword":"'+key_word+'","conditions":{"excep_tab":"0","ill_tab":"0","area":"0","cStatus":"0","xzxk":"0","xzcf":"0","dydj":"0"},"sourceType":"I"}')
    data_ = json.dumps(body,ensure_ascii=False)  #控制转json后的乱码
    headers = {
        'Content-Type':'application/raw',
        'Host':'app.gsxt.gov.cn',
        'Cookie':'JSESSIONID=D26788EACA25EEF8BFAA4091585D7C77; SECTOKEN=7184035096920064216; __jsluid=da5c7271631d61c4fb14a75350f62e48',
        'Connection':'keep-alive',
        'Accept':'application/json',
        'User-Agent':'Mozilla/5.0 (iPad; CPU OS 12_1_3 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/16D39 Html5Plus/1.0',
        'Accept-Language':'zh-cn',
        'Accept-Encoding':'br, gzip, deflate',
        'X-Requested-With':'XMLHttpRequest'
    }
    response = requests.post(url, data=data_.encode('utf-8'), headers=headers,verify=False)
    # 返回信息
    versionInfo = response.text
    versionInfoPython = json.loads(versionInfo)
    dataList = versionInfoPython['data']['result']
    print(dataList['data'][0])

    from pymongo import MongoClient

    conn = MongoClient('127.0.0.1', 27017)
    db = conn.Test  #连接mydb数据库，没有则自动创建
    my_set = db.gsxy
    my_set.insert(dataList['data'][0])
    # 返回响应头
    # print(response.status_code)

i = 0
l = 1
def url_pa(i,l):
    # excel文件名称
    wb = load_workbook('shuaizhen.xlsx')
    sheet = wb.worksheets[0]
    lou_name = []
    for rows in range(i, l):
        key_word = sheet.cell(row=rows + 1, column=1).value
        lou_name.append(key_word)
        print(key_word, rows)
        print('正在搜索，请稍后')
        try:
            s1 = craw(key_word)
        except Exception as err:
            print("err!!!")
            print(key_word)
            ws1.append([key_word])
            wb1.save('QYXX.xlsx')
    i = i+1
    l = l+1
    threading.Timer(3, url_pa, [i, l]).start()

if __name__ == '__main__':
    threading.Timer(1,url_pa, [i,l]).start()