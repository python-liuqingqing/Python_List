# coding=utf-8
import time
import datetime
import mysql.connector
import re
import redis
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from selenium import webdriver
import mysql.connector
driver = webdriver.Chrome()
driver.get('https://www.baidu.com')
i = 0
l = 1

def url_pa(i,l):
    wb = load_workbook('gjc.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    for rows in range(i, l):
        key_word = sheet.cell(row=rows + 1, column=2).value
        print(key_word, rows)
        print('正在搜索，请稍后')
        # main(key_word)
        js = 'window.open("http://www.360doc.com/search.html?type=0&word=案例、' + key_word + '、企业所得税处理、填报实务");'
        driver.execute_script(js)
        handles = driver.window_handles  # 获取当前窗口句柄集合（列表类型）
        for handle in handles:  # 切换窗口（切换到搜狗）
            if handle != driver.current_window_handle:
                driver.switch_to.window(handle)
        time.sleep(5)
        html = driver.find_element_by_id('divsearchresult').get_attribute('innerHTML')
        soup = BeautifulSoup(html, 'lxml')  # 对html进行解析
        href_ = soup.select('span a')
        for line in href_:
            print(line['href']+line.text)

            db = mysql.connector.connect(user='root', password='123456', host='192.168.1.130', database='master')
            cursor = db.cursor()

            insert_color = (
                "INSERT INTO 360doc(href,name,class)" "VALUES(%s,%s,%s)")
            data_color = (line['href'],line.text,key_word)
            cursor.execute(insert_color, data_color)
            db.commit()


        driver.close()
        driver.switch_to_window(handles[0])

    i = i+1
    l = l+1
    url_pa(i, l)

url_pa(i,l)